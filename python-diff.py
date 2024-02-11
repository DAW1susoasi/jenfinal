#!/usr/bin/python3

import os
import sys
import openpyxl

print("Welcome to the Diff Excel User System")

vCampos=["Login","Nombre","Telefono","Correo"]

ruta_ayer = sys.argv[1]
ruta_hoy = sys.argv[2]

print("Viejo fue: " + str(ruta_ayer))
print("Nuevo fue: " + str(ruta_hoy))

wb_ayer = openpyxl.load_workbook(ruta_ayer)
ws_ayer = wb_ayer.active

wb_hoy = openpyxl.load_workbook(ruta_hoy)
ws_hoy = wb_hoy.active

if os.path.exists("meta-script.sh"):
    print("Meta Script Cleaning System ")
    os.remove("meta-script.sh")
    
meta_script = open("meta-script.sh",'x')
meta_script.write("#!/bin/bash\n\n")
meta_script.write("# DO NOT EDIT THIS SCRIPT \n")
meta_script.write("# IT WILL BE AUTOMAGICALLY GENERATED\n\n")
meta_script.write("# Orig "+ruta_ayer+"\n")
meta_script.write("# Dest "+ruta_hoy+"\n")

def insertarModificar():
    aux_id_hoy=ws_hoy.cell(row=1,column=1).value
    fila_hoy_procesada = 1
    meta_script.write("\n# NUEVOS CONTRATOS/MODIFICACIONES \n")
    while (aux_id_hoy != None):
        fila_ayer_procesada = 1
        aux_id_ayer = ws_ayer.cell(row=1,column=1).value
        vCambiosUsuario = []
        encontrado = False
        # Leo las filas de ayer mientas el id de ayer sea distinto al id de hoy
        while (aux_id_ayer != None and not encontrado):
            if aux_id_ayer == aux_id_hoy:
                # Paro de recorrer el bucle interno
                encontrado = True
                hayCambios = False
                # del 2 al 5 incluidos
                for campo in range(2,6):
                    if ws_ayer.cell(row=fila_ayer_procesada,column=campo).value != ws_hoy.cell(row=fila_hoy_procesada,column=campo).value:
                        # Meto en un array las posiciones de los campos que han cambiado
                        vCambiosUsuario.append(campo-2)
                        hayCambios = True
                
                if hayCambios:
                    print(" * " + str(ws_ayer.cell(row=fila_ayer_procesada,column=1).value) + " ha cambiado:")
                    #print(vCambiosUsuario)
                    # Lo último que se ha de modificar es el login
                    for campoCambiado in range(len(vCambiosUsuario)-1, -1, -1):
                        #print(vCambiosUsuario[campoCambiado])
                        auxC = int(vCambiosUsuario[campoCambiado])+2
                        if vCambiosUsuario[campoCambiado] == 0:
                            print("   " + vCampos[vCambiosUsuario[campoCambiado]] + ": "+ str(ws_ayer.cell(row=fila_ayer_procesada,column=auxC).value) + " -> " + str(ws_hoy.cell(row=fila_hoy_procesada,column=auxC).value))
                            meta_script.write("# Modifico login de " + str(ws_ayer.cell(row=fila_ayer_procesada,column=2).value) +"\n")
                            meta_script.write("usermod -l "+str(ws_hoy.cell(row=fila_hoy_procesada,column=auxC).value)+" "+str(ws_ayer.cell(row=fila_ayer_procesada,column=2).value)+"\n")
                        if vCambiosUsuario[campoCambiado] == 1:
                            print("   " + vCampos[vCambiosUsuario[campoCambiado]] + ": "+ str(ws_ayer.cell(row=fila_ayer_procesada,column=auxC).value) + " -> " + str(ws_hoy.cell(row=fila_hoy_procesada,column=auxC).value))
                            meta_script.write("# Modifico nombre completo de " + str(ws_ayer.cell(row=fila_ayer_procesada,column=2).value) +"\n")
                            meta_script.write("chfn -f \""+str(ws_hoy.cell(row=fila_hoy_procesada,column=auxC).value)+"\" "+str(ws_ayer.cell(row=fila_ayer_procesada,column=2).value)+"\n")
                        if vCambiosUsuario[campoCambiado] == 2:
                            print("   " + vCampos[vCambiosUsuario[campoCambiado]] + ": "+ str(ws_ayer.cell(row=fila_ayer_procesada,column=auxC).value) + " -> " + str(ws_hoy.cell(row=fila_hoy_procesada,column=auxC).value))
                            meta_script.write("# Modifico teléfono de " + str(ws_ayer.cell(row=fila_ayer_procesada,column=2).value) +"\n")
                            meta_script.write("chfn -h \""+str(ws_hoy.cell(row=fila_hoy_procesada,column=auxC).value)+"\" "+str(ws_ayer.cell(row=fila_ayer_procesada,column=2).value)+"\n")
                        if vCambiosUsuario[campoCambiado] == 3:
                            print("   " + vCampos[vCambiosUsuario[campoCambiado]] + ": "+ str(ws_ayer.cell(row=fila_ayer_procesada,column=auxC).value) + " -> " + str(ws_hoy.cell(row=fila_hoy_procesada,column=auxC).value))
                            meta_script.write("# Modifico correo de " + str(ws_ayer.cell(row=fila_ayer_procesada,column=2).value) +"\n")
                            meta_script.write("chfn -o \""+str(ws_hoy.cell(row=fila_hoy_procesada,column=auxC).value)+"\" "+str(ws_ayer.cell(row=fila_ayer_procesada,column=2).value)+"\n")

            fila_ayer_procesada = fila_ayer_procesada + 1
            aux_id_ayer = ws_ayer.cell(row=fila_ayer_procesada,column=1).value
        # Si es un nuevo contrato 
        if not encontrado:
            auxUser = ws_hoy.cell(row=fila_hoy_procesada,column=2).value
            auxFull = ws_hoy.cell(row=fila_hoy_procesada,column=3).value
            auxTel = ws_hoy.cell(row=fila_hoy_procesada,column=4).value
            auxMail = ws_hoy.cell(row=fila_hoy_procesada,column=5).value
            auxUID = ws_hoy.cell(row=fila_hoy_procesada,column=1).value
            print(" * Contrato a " + str(ws_hoy.cell(row=fila_hoy_procesada,column=1).value))
            meta_script.write("# Adding " + ws_hoy.cell(row=fila_hoy_procesada,column=3).value+"\n")
            meta_script.write("useradd -m -d \"/home/"+auxUser+"\" -s \"/bin/bash\" -u "+str(auxUID)+" -c \""+auxFull+", ,"+str(auxTel)+", ,"+auxMail+"\" \""+auxUser+"\"\n" )
            meta_script.write("echo \""+auxUser+":"+str(auxTel)+"\"| chpasswd \n\n")

        # Paso a una nueva fila del bucle externo
        fila_hoy_procesada = fila_hoy_procesada + 1
        aux_id_hoy = ws_hoy.cell(row=fila_hoy_procesada,column=1).value

def eliminar():
    # Busca en el archivo de ayer los que no estén en el archivo de hoy
    meta_script.write("\n\n# DESPIDOS PROCEDENTES :  \n")

    aux_id_ayer=ws_ayer.cell(row=1,column=1).value
    fila_ayer_procesada = 1
    # Leo todas las filas de ayer
    while (aux_id_ayer != None):
        
        aux_id_hoy=ws_hoy.cell(row=1,column=1).value
        fila_hoy_procesada = 1
        encontrado = False
        # Leo las filas de hoy mientas el id de ayer sea distinto al id de hoy
        while (aux_id_hoy != None and not encontrado):
            # Si el id de ayer es igual al id de hoy 
            if aux_id_hoy == ws_ayer.cell(row=fila_ayer_procesada,column=1).value:
                # No es un despido y paro de recorrer el bucle interno
                encontrado = True
            # Paso a una nueva fila del bucle interno
            fila_hoy_procesada = fila_hoy_procesada + 1
            aux_id_hoy = ws_hoy.cell(row=fila_hoy_procesada,column=1).value
        # Si s un despido 
        if not encontrado:
            print(" * Despide a " + str(ws_ayer.cell(row=fila_ayer_procesada,column=1).value))
            meta_script.write("# Deleting "+ws_ayer.cell(row=fila_ayer_procesada,column=3).value+"\n")
            meta_script.write("deluser "+ws_ayer.cell(row=fila_ayer_procesada,column=2).value+"\n")
            
        # Paso a una nueva fila del bucle externo
        fila_ayer_procesada = fila_ayer_procesada + 1
        aux_id_ayer = ws_ayer.cell(row=fila_ayer_procesada,column=1).value

eliminar()      
insertarModificar()    

# Print exit 0 
meta_script.write("exit 0\n")
meta_script.close()

sys.exit(0)





